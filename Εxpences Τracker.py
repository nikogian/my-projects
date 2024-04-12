import os
from datetime import datetime
import openpyxl


# Excel output function

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def save_expenses(obj,value,date,path):
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

        headers = ['Date', 'Object', 'Value', 'Sum']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = header
            sheet[f'{col_letter}1'].font = Font(bold=True)

    total_sum = value
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row-1): 
        for cell in row:
            if cell.column == openpyxl.utils.column_index_from_string('C'):
                total_sum += cell.value 
    
    row_num = sheet.max_row 
    if date == 'Y':
        sheet[f'A{row_num}'] = datetime.now().strftime('%d-%m-%Y')
    else:
        sheet[f'A{row_num}'] = input('Type the date of expense (format dd-mm-yyyy): ')

    sheet[f'B{row_num}'] = obj
    sheet[f'B{row_num+1}'] = 'sum'
    #sheet[f'B{row_num+1}'].font = Font(bold=True)
    sheet[f'C{row_num}'] = value
    sheet[f'C{row_num+1}'] = total_sum
    #sheet[f'C{row_num+1}'].font = Font(bold=True)
    sheet[f'D{row_num}'] = input('Type extra details: ')
    sheet[f'D{row_num+1}'] = '€'
    #sheet[f'D{row_num+1}'].font = Font(bold=True)
    
    workbook.save(path)


# Operation

add_expense = 'Y'

while add_expense == 'Y':

#user's input for object and value
    exp_obj = input("Add expense's object: ").upper()

    while True:
        try:
            exp_value = float(input('Add expense (€): '))
            break
        except ValueError:
            print('Wrong input, please type a valid amount of money')

    print(exp_obj, exp_value , '€\n')


# Save

    auto_date = input('Press ENTER to use current date or type "man" to add manual date" ').upper() or "Y"
#'my_path' should be modified to the actual path on your device
    save_path = 'my_path'

    save_expenses(exp_obj,exp_value,auto_date,save_path)

    add_expense = input('Press ENTER to Exit or type "y" to add more expenses: ').upper() or 'N'


#dictionary (optional)
#exp_dict = {exp_obj:exp_value}
#print(exp_dict)
    
