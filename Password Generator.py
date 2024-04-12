import os
import time
import random
import string

# Excel output

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def save_credentials(acc,user,psw,extra,path):
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

        headers = ['Account', 'Username', 'Password', 'Extra']
        for x, y in enumerate(headers, 1):
            col_letter = get_column_letter(x)
            sheet[f"{col_letter}1"] = y
            sheet[f"{col_letter}1"].font = Font(bold=True)

    row_num = sheet.max_row + 1
    sheet[f"A{row_num}"] = acc
    sheet[f"B{row_num}"] = user
    sheet[f"C{row_num}"] = psw
    sheet[f"D{row_num}"] = extra

    workbook.save(path)


# Shortcuts function can be used to automatically fill your corresponding email address by simply typing gmail, hotmail or yahoo
def shortcut(var):
    shorts = {"gmail":"your_address@gmail.com","hotmail":"your_address@hotmail.com",
"yahoo":"your_address@yahoo.com"}
    if var in shorts.keys():
        var = shorts[var]
    return var


# Password generation function
def psw_gen():
    test = input("Press ENTER for RANDOM password or type 'man' for manual password: ") or "N"
    if test != "N":
        check = "N"
        while check == "N":
            password = input("Enter password: ")
            check = input(f"Type 'y' if this is the right password: {password} , or press Enter to try again: ").upper() or "N"
    else:
        while True:
            length = input("Password length (press Enter for default value 12): ") or "12"
            try:
                length = int(length)
                break
            except ValueError:
                print("Invalid input, please enter a valid number")
        characters = string.ascii_letters + string.digits + string.punctuation
        password = "".join(random.choices(characters, k=length))
    return password


# Loop
while True:

#account reference
    account = input("Account: ").upper()
    username = input("Username or email: ")
    username = shortcut(username)
    password = psw_gen()
 
    print(f"\nAccount: {account}")
    print(f"Username: {username}")
    print(f"Password: {password}")

    check = input("Type 'res' if there is a mistake above, or press ENTER to continue: ")
    if check == "res":
        continue

    details = input("Type extra details: ")
    details = shortcut(details)


#save
    save = input("Press ENTER to save the credentials or type 'n' to skip: ").upper() or "Y"
    
    if save == "Y":
# Modify the path to correspond to the actual path on your device
        file_path = "my_path/excelfile.xlsx"
        print(f"Saving the credentials on:{file_path}")
        print("...")
        save_credentials(account,username,password,details,file_path)
        time.sleep(1)
    else:
        print("Saving skipped...")
        time.sleep(1)

    print("Proccess done!")
    new = input("Type 'y' to add more credentials, or press ENTER to exit: ").upper()
    if new != "Y":
        break
    print("")

print("Work completed")
time.sleep(1)


            
